"""
01_konsolidierung.py — PSBs + SAP einlesen, validieren, konsolidiertes Excel exportieren.

Ausfuehrung: python 01_konsolidierung.py
Ergebnis:    output/review/konsolidiert.xlsx
"""

import sys
from pathlib import Path
from datetime import datetime, date
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pandas as pd
import numpy as np

from config import (
    PSB_DIR, DASHBOARD_FILE, SAP_FILE,
    OUTPUT_REVIEW,
    AMPEL_NORM, PAG_NORM, PHASE_NORM,
    PSB_CELLS, PSB_MEILENSTEIN_ROWS, PSB_MEILENSTEIN_COLS,
    PSB_TEXT_RANGES, PSB_ZIELWERT_CELLS,
    DASHBOARD_COLS, FEIERTAGE, AMPEL_FARBEN, STICHTAG, QUARTAL,
)


# =============================================================================
# 1. PSB-Reader
# =============================================================================

def _get_merged_value(ws, cell_ref: str) -> Any:
    """Liest den Wert einer Zelle, auch wenn sie Teil einer Merged-Region ist."""
    cell = ws[cell_ref]
    if cell.value is not None:
        return cell.value
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left.value
    return None


def _read_text_block(ws, start_ref: str, end_ref: str) -> str:
    """Liest einen Merged-Textblock und gibt den zusammengefuegten Text zurueck."""
    start_cell = ws[start_ref]
    start_row, start_col = start_cell.row, start_cell.column
    end_cell = ws[end_ref]
    end_row, end_col = end_cell.row, end_cell.column

    texts = []
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                texts.append(str(val).strip())
    # Auch merged-cell Wert pruefen
    merged_val = _get_merged_value(ws, start_ref)
    if merged_val and str(merged_val).strip() not in texts:
        texts.insert(0, str(merged_val).strip())
    return "\n".join(texts) if texts else ""


def _normalize_date(val: Any) -> Any:
    """Versucht ein Datum zu normalisieren. Gibt date, None oder den Originalwert zurueck."""
    if val is None:
        return None
    if isinstance(val, datetime):
        d = val.date()
        if d.year < 2000:  # Falsches Datum (z.B. 1905)
            return None
        return d
    if isinstance(val, date):
        if val.year < 2000:
            return None
        return val
    s = str(val).strip()
    if not s or s == "TT.MM.JJJJ" or s == "–" or s == "-":
        return None
    # Quartalsformat "Q1 26" → letzter Tag des Quartals
    if s.upper().startswith("Q") and len(s) >= 4:
        try:
            parts = s.upper().replace("Q", "").strip().split()
            q = int(parts[0])
            year = int(parts[1]) if len(parts) > 1 else None
            if year and year < 100:
                year += 2000
            if year and 1 <= q <= 4:
                month = q * 3
                import calendar
                day = calendar.monthrange(year, month)[1]
                return date(year, month, day)
        except (ValueError, IndexError):
            pass
    # DD.MM.YYYY
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return s  # Originalwert zurueckgeben wenn nicht parsebar


def read_psb(filepath: Path) -> dict:
    """Liest einen einzelnen PSB ein und gibt ein Dictionary mit allen Feldern zurueck."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Statusbericht"]

    record = {}

    # Identifikation
    for key, cell_ref in PSB_CELLS.items():
        record[key] = _get_merged_value(ws, cell_ref)

    # Normalisierungen
    if record.get("ampelstatus"):
        record["ampelstatus"] = AMPEL_NORM.get(str(record["ampelstatus"]).strip(), record["ampelstatus"])
    if record.get("pag"):
        record["pag"] = PAG_NORM.get(str(record["pag"]).strip(), record["pag"])
    if record.get("projektphase"):
        record["projektphase"] = PHASE_NORM.get(str(record["projektphase"]).strip(), record["projektphase"])

    # Datumsfelder
    record["berichtsdatum"] = _normalize_date(record.get("berichtsdatum"))
    record["projektende"] = _normalize_date(record.get("projektende"))

    # LV-Nr: Prefix "LV25-27 " entfernen
    lv_nr = record.get("lv_nr", "")
    if lv_nr and isinstance(lv_nr, str):
        import re
        match = re.search(r"([A-D]\d+(?:\.\d+)*)", lv_nr)
        if match:
            record["lv_nr"] = match.group(1)

    # Fertigstellungsgrad
    fg = record.get("fertigstellungsgrad")
    if fg is not None:
        try:
            record["fertigstellungsgrad"] = float(fg)
        except (ValueError, TypeError):
            record["fertigstellungsgrad"] = 0.0
    else:
        record["fertigstellungsgrad"] = 0.0

    # Meilensteine
    meilensteine = []
    for row_num in PSB_MEILENSTEIN_ROWS:
        bez = ws[f"{PSB_MEILENSTEIN_COLS['bezeichnung']}{row_num}"].value
        if bez and str(bez).strip():
            ms = {
                "bezeichnung": str(bez).strip(),
                "plan_lv": _normalize_date(ws[f"{PSB_MEILENSTEIN_COLS['plan_lv']}{row_num}"].value),
                "plan_aktuell": _normalize_date(ws[f"{PSB_MEILENSTEIN_COLS['plan_aktuell']}{row_num}"].value),
                "ist": _normalize_date(ws[f"{PSB_MEILENSTEIN_COLS['ist']}{row_num}"].value),
                "erlaeuterung": ws[f"{PSB_MEILENSTEIN_COLS['erlaeuterung']}{row_num}"].value,
            }
            meilensteine.append(ms)
    record["meilensteine"] = meilensteine
    record["meilensteine_text"] = "; ".join(m["bezeichnung"] for m in meilensteine)
    record["meilensteine_anzahl"] = len(meilensteine)

    # Textbloecke
    for key, (start, end) in PSB_TEXT_RANGES.items():
        record[key] = _read_text_block(ws, start, end)

    # Zielwerte
    zielwerte = {}
    for key, cell_ref in PSB_ZIELWERT_CELLS.items():
        zielwerte[key] = _get_merged_value(ws, cell_ref)
    record["zielwerte"] = zielwerte

    # Start-Datum: erstes Meilenstein-Plan-Datum
    start_dates = [m["plan_lv"] for m in meilensteine if isinstance(m["plan_lv"], date)]
    record["start"] = min(start_dates) if start_dates else None

    # Ende-Datum: aus Projektende-Feld oder letzter Meilenstein
    ende = record.get("projektende")
    if isinstance(ende, date):
        record["ende"] = ende
    else:
        end_dates = [m["plan_lv"] for m in meilensteine if isinstance(m["plan_lv"], date)]
        record["ende"] = max(end_dates) if end_dates else None

    wb.close()
    return record


def read_all_psbs() -> list[dict]:
    """Liest alle PSB-Dateien im PSB_DIR ein."""
    psbs = []
    for f in sorted(PSB_DIR.glob("PSB_*.xlsx")):
        if f.name.startswith("~$"):
            continue
        print(f"  Lese PSB: {f.name}")
        psb = read_psb(f)
        psb["_quelldatei"] = f.name
        psbs.append(psb)
    return psbs


# =============================================================================
# 2. SAP-Reader
# =============================================================================

def read_sap() -> pd.DataFrame:
    """Liest SAP-Finanzdaten ein. Betraege werden negiert (Ausgaben → positiv)."""
    wb = openpyxl.load_workbook(SAP_FILE, data_only=True)
    ws = wb.active

    records = []
    for row in ws.iter_rows(min_row=5, values_only=False):
        vals = [c.value for c in row]
        if not vals[0]:
            continue
        le_nummer = str(vals[0]).strip()
        objekt_text = str(vals[1]).strip() if vals[1] else ""

        # LV-Nrn aus Objekt-Text extrahieren (kann mehrere enthalten)
        lv_nrs = _extract_lv_nrs(objekt_text)

        def neg(v):
            """Negiert Betraege (SAP: negativ = Ausgabe)."""
            if v is None:
                return 0.0
            try:
                return abs(float(v))
            except (ValueError, TypeError):
                return 0.0

        base = {
            "le_nummer": le_nummer,
            "objekt_text": objekt_text,
            "kostenstelle": le_nummer,
            "plankosten_2025": neg(vals[3]),
            "istkosten_2025": neg(vals[4]),
            "plankosten_2026": neg(vals[5]),
            "istkosten_2026": neg(vals[6]),
            "plankosten_2027": neg(vals[7]),
        }
        # Eine Zeile pro LV-Nr (bei Mehrfach-Zuordnung wird Budget geteilt)
        for lv_nr in (lv_nrs or [objekt_text]):
            rec = {**base, "lv_nr_sap": lv_nr}
            if len(lv_nrs) > 1:
                for k in ["plankosten_2025", "istkosten_2025", "plankosten_2026",
                           "istkosten_2026", "plankosten_2027"]:
                    rec[k] = base[k] / len(lv_nrs)
            records.append(rec)

    wb.close()

    df = pd.DataFrame(records)
    # Plankosten Gesamt
    df["plankosten_gesamt"] = df["plankosten_2025"] + df["plankosten_2026"] + df["plankosten_2027"]
    # Istkosten 2027 noch nicht vorhanden
    df["istkosten_2027"] = 0.0
    df["istkosten_gesamt"] = df["istkosten_2025"] + df["istkosten_2026"] + df["istkosten_2027"]
    df["istkosten_prozent"] = np.where(
        df["plankosten_gesamt"] > 0,
        df["istkosten_gesamt"] / df["plankosten_gesamt"],
        0.0,
    )
    return df


def _extract_lv_nrs(text: str) -> list[str]:
    """Extrahiert alle LV-Nummern aus dem SAP-Objekttext.
    z.B. 'D3.3.2+D3.3.3 Bauvorhaben' → ['D3.3.2', 'D3.3.3']
    z.B. 'A1.1.1.1.1 Labor' → ['A1.1.1.1.1']"""
    import re
    return re.findall(r"[A-D]\d+(?:\.\d+)*", text)


# =============================================================================
# 3. Dashboard-Reader (bestehendes Dashboard)
# =============================================================================

def read_dashboard() -> pd.DataFrame:
    """Liest das bestehende Dashboard (Portfolio_Daten) ein."""
    wb = openpyxl.load_workbook(DASHBOARD_FILE, data_only=True)
    ws = wb["Portfolio_Daten"]

    records = []
    for row in ws.iter_rows(min_row=5, max_row=122, values_only=False):
        vals = {col_idx: row[col_idx - 1].value for col_idx in DASHBOARD_COLS}
        if not vals.get(2) and not vals.get(5):
            continue
        record = {name: vals[col_idx] for col_idx, name in DASHBOARD_COLS.items()}
        records.append(record)

    wb.close()
    return pd.DataFrame(records)


def read_dashboard_ppm_fields() -> pd.DataFrame:
    """Liest nur die PPM-manuellen Felder aus dem bestehenden Dashboard,
    damit diese bei der Konsolidierung uebernommen werden koennen."""
    df = read_dashboard()
    ppm_cols = [
        "lv_nummer", "leistungsbereich", "kapitelzuordnung", "lv_periode",
        "projektauftrag", "psb_vorhanden", "lenkungsausschuss",
        "externe_kooperationspartner", "thema_begleitgespraech",
        "budgeteinbehalt", "prioritaet",
    ]
    return df[[c for c in ppm_cols if c in df.columns]]


# =============================================================================
# 4. Konsolidierung
# =============================================================================

def networkdays(start_date, end_date, holidays=None):
    """Berechnet Werktage zwischen zwei Daten (analog NETWORKDAYS in Excel)."""
    if not isinstance(start_date, date) or not isinstance(end_date, date):
        return None
    if holidays is None:
        holidays = FEIERTAGE
    holiday_set = set(holidays)
    count = 0
    current = start_date
    from datetime import timedelta
    while current <= end_date:
        if current.weekday() < 5 and current not in holiday_set:
            count += 1
        current += timedelta(days=1)
    return count


def konsolidiere(psbs: list[dict], sap_df: pd.DataFrame, ppm_df: pd.DataFrame) -> pd.DataFrame:
    """Fuehrt PSBs, SAP-Daten und PPM-Felder zu einem DataFrame zusammen."""
    rows = []
    for psb in psbs:
        lv_nr = str(psb.get("lv_nr", "")).strip()

        # SAP-Daten per LV-Nr joinen (exact oder startswith)
        sap_match = sap_df[sap_df["lv_nr_sap"] == lv_nr]
        if sap_match.empty:
            sap_match = sap_df[sap_df["lv_nr_sap"].str.startswith(lv_nr, na=False)]
        sap_row = sap_match.iloc[0].to_dict() if len(sap_match) > 0 else {}

        # PPM-Felder per LV-Nr joinen
        ppm_match = ppm_df[ppm_df["lv_nummer"] == lv_nr]
        ppm_row = ppm_match.iloc[0].to_dict() if len(ppm_match) > 0 else {}

        # Dauer berechnen
        start = psb.get("start")
        ende = psb.get("ende")
        dauer = networkdays(start, ende) if start and ende else None

        # Berechnete Kostenfelder
        istkosten_gesamt = (
            sap_row.get("istkosten_2025", 0) +
            sap_row.get("istkosten_2026", 0) +
            sap_row.get("istkosten_2027", 0)
        )
        plankosten_gesamt = sap_row.get("plankosten_gesamt", 0)
        istkosten_prozent = istkosten_gesamt / plankosten_gesamt if plankosten_gesamt else 0.0

        row = {
            "lv_nummer": lv_nr,
            "leistungsbereich": ppm_row.get("leistungsbereich", ""),
            "kapitelzuordnung": ppm_row.get("kapitelzuordnung", ""),
            "projektname": psb.get("bezeichnung", ""),
            "kurzbeschreibung": psb.get("kurzbeschreibung", ""),
            "meilensteine": psb.get("meilensteine_text", ""),
            "projektleitung": psb.get("projektleitung", ""),
            "pag": psb.get("pag", ""),
            "organisationseinheiten": psb.get("organisationseinheiten", ""),
            "lv_periode": ppm_row.get("lv_periode", "LV25-27"),
            "projektauftrag": ppm_row.get("projektauftrag", ""),
            "psb_vorhanden": "Ja",
            "lenkungsausschuss": ppm_row.get("lenkungsausschuss", ""),
            "externe_kooperationspartner": ppm_row.get("externe_kooperationspartner", ""),
            "thema_begleitgespraech": ppm_row.get("thema_begleitgespraech", ""),
            "budgeteinbehalt": ppm_row.get("budgeteinbehalt", ""),
            "prioritaet": ppm_row.get("prioritaet", ""),
            "risiko": psb.get("risikoeinschaetzung", ""),
            "start": start,
            "ende": ende,
            "dauer_werktage": dauer,
            "kostenstelle": sap_row.get("kostenstelle", ""),
            "plankosten_2025": sap_row.get("plankosten_2025", 0),
            "plankosten_2026": sap_row.get("plankosten_2026", 0),
            "plankosten_2027": sap_row.get("plankosten_2027", 0),
            "plankosten_gesamt": plankosten_gesamt,
            "istkosten_2025": sap_row.get("istkosten_2025", 0),
            "istkosten_2026": sap_row.get("istkosten_2026", 0),
            "istkosten_2027": sap_row.get("istkosten_2027", 0),
            "istkosten_gesamt": istkosten_gesamt,
            "istkosten_prozent": istkosten_prozent,
            "projektphase": psb.get("projektphase", ""),
            "ampelstatus": psb.get("ampelstatus", ""),
            "fertigstellungsgrad": psb.get("fertigstellungsgrad", 0),
            "ziele": psb.get("zielwerte", {}).get("ziel_text", ""),
            "status_aktuell": psb.get("erlaeuterung_ampel", ""),
            "risiken_aktuell": psb.get("risiken", ""),
            "entscheidung_aktuell": psb.get("entscheidungsbedarf", ""),
            "status_vorperiode": "",
            "risiken_vorperiode": "",
            "entscheidung_vorperiode": "",
            "_quelldatei": psb.get("_quelldatei", ""),
        }
        rows.append(row)

    return pd.DataFrame(rows)


# =============================================================================
# 5. Validierung
# =============================================================================

def validiere(df: pd.DataFrame) -> pd.DataFrame:
    """Prueft jeden Datensatz und fuegt eine Validierungsspalte hinzu."""
    fehler = []
    for idx, row in df.iterrows():
        probleme = []
        # Pflichtfelder
        if not row.get("lv_nummer"):
            probleme.append("LV-Nummer fehlt")
        if not row.get("projektname"):
            probleme.append("Projektname fehlt")
        if not row.get("pag"):
            probleme.append("PAG fehlt")
        if not row.get("ampelstatus"):
            probleme.append("Ampelstatus fehlt")
        elif row["ampelstatus"] not in ("In Ordnung", "Vorsicht", "Krise"):
            probleme.append(f"Ungueltiger Ampelstatus: {row['ampelstatus']}")
        if not row.get("projektphase"):
            probleme.append("Projektphase fehlt")

        # Datumspruefungen
        if row.get("start") and isinstance(row["start"], str):
            probleme.append(f"Start nicht als Datum parsebar: {row['start']}")
        if row.get("ende") and isinstance(row["ende"], str):
            probleme.append(f"Ende nicht als Datum parsebar: {row['ende']}")

        # Fertigstellungsgrad
        fg = row.get("fertigstellungsgrad", 0)
        if fg and not (0 <= fg <= 1):
            probleme.append(f"Fertigstellungsgrad ausserhalb 0-1: {fg}")

        fehler.append("; ".join(probleme) if probleme else "OK")

    df["_validierung"] = fehler
    return df


# =============================================================================
# 6. Excel-Export
# =============================================================================

def export_review_excel(df: pd.DataFrame, output_path: Path):
    """Exportiert das konsolidierte DataFrame als formatiertes Excel."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Spalten ohne interne Felder (ausser _validierung)
        export_cols = [c for c in df.columns if not c.startswith("_") or c == "_validierung"]
        df[export_cols].to_excel(writer, sheet_name="Konsolidiert", index=False)

        ws = writer.sheets["Konsolidiert"]

        # Formatierung
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Header formatieren
        for col_idx, col_name in enumerate(export_cols, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        # Ampelstatus einfaerben
        ampel_col = export_cols.index("ampelstatus") + 1 if "ampelstatus" in export_cols else None
        validierung_col = export_cols.index("_validierung") + 1 if "_validierung" in export_cols else None

        fills = {
            "In Ordnung": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "Vorsicht": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "Krise": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        }

        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(export_cols) + 1):
                ws.cell(row=row_idx, column=col_idx).border = thin_border

            # Ampelfarbe
            if ampel_col:
                ampel_val = ws.cell(row=row_idx, column=ampel_col).value
                if ampel_val in fills:
                    ws.cell(row=row_idx, column=ampel_col).fill = fills[ampel_val]

            # Validierungsspalte rot wenn nicht OK
            if validierung_col:
                val_cell = ws.cell(row=row_idx, column=validierung_col)
                if val_cell.value and val_cell.value != "OK":
                    val_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    val_cell.font = Font(color="9C0006")

        # Spaltenbreiten
        for col_idx in range(1, len(export_cols) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # Autofilter
        ws.auto_filter.ref = ws.dimensions

    print(f"  Exportiert: {output_path}")


# =============================================================================
# Main
# =============================================================================

def main():
    print(f"=== Konsolidierung Projektportfolio LV-Vorhaben ({QUARTAL}) ===\n")

    # 1. PSBs lesen
    print("1. PSBs einlesen...")
    psbs = read_all_psbs()
    print(f"   {len(psbs)} PSBs gelesen.\n")

    # Verifikation: Kernwerte ausgeben
    for psb in psbs:
        lv = psb.get("lv_nr", "?")
        print(f"   {lv}: Ampel={psb.get('ampelstatus')}, Phase={psb.get('projektphase')}, "
              f"Fortschritt={psb.get('fertigstellungsgrad')}, Risiko={psb.get('risikoeinschaetzung')}, "
              f"Meilensteine={psb.get('meilensteine_anzahl')}")
    print()

    # 2. SAP lesen
    print("2. SAP-Finanzdaten einlesen...")
    sap_df = read_sap()
    print(f"   {len(sap_df)} SAP-Eintraege gelesen.\n")

    # 3. Bestehendes Dashboard lesen (PPM-Felder)
    print("3. Bestehendes Dashboard einlesen (PPM-Felder)...")
    ppm_df = read_dashboard_ppm_fields()
    print(f"   {len(ppm_df)} Dashboard-Eintraege gelesen.\n")

    # 4. Konsolidieren
    print("4. Konsolidierung...")
    df = konsolidiere(psbs, sap_df, ppm_df)
    print(f"   {len(df)} konsolidierte Eintraege.\n")

    # 5. Validieren
    print("5. Validierung...")
    df = validiere(df)
    n_ok = (df["_validierung"] == "OK").sum()
    n_fehler = len(df) - n_ok
    print(f"   {n_ok} OK, {n_fehler} mit Problemen.")
    if n_fehler > 0:
        for _, row in df[df["_validierung"] != "OK"].iterrows():
            print(f"   ! {row['lv_nummer']}: {row['_validierung']}")
    print()

    # 6. Excel exportieren
    print("6. Excel-Export...")
    output_path = OUTPUT_REVIEW / "konsolidiert.xlsx"
    export_review_excel(df, output_path)

    print(f"\nFertig. Ergebnis: {output_path}")
    return df


if __name__ == "__main__":
    main()
