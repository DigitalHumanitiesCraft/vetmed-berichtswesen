"""
check.py — Verifikation der konsolidierten Daten gegen Quelldateien und Regeln.

Ausfuehrung: python check.py
Voraussetzung: python 01_konsolidierung.py muss vorher gelaufen sein.

Prueft 6 Kategorien:
  1. Quelldaten-Abgleich (bekannte Sollwerte pro PSB)
  2. Normalisierung (Ampel, PAG, Phase)
  3. Datumsfelder (kein Jahr < 2000, keine Platzhalter)
  4. Budget/SAP (keine negativen Betraege, keine Division by Zero)
  5. Vollstaendigkeit (Pflichtfelder, Projekt-Anzahl)
  6. Validierungsspalte (_validierung vorhanden)
"""

import sys
from pathlib import Path
from datetime import datetime, date

import openpyxl
import pandas as pd

from config import (
    OUTPUT_REVIEW, PSB_DIR, ERWARTETE_WERTE,
    AMPEL_NORM, PAG_NORM, PHASE_NORM,
)


def lade_konsolidiert(pfad: Path) -> pd.DataFrame:
    """Liest konsolidiert.xlsx als DataFrame."""
    wb = openpyxl.load_workbook(pfad, data_only=True)
    ws = wb["Konsolidiert"]

    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val).strip() if val else f"col_{col}")

    rows = []
    for row in range(2, ws.max_row + 1):
        vals = {}
        for col_idx, header in enumerate(headers, 1):
            vals[header] = ws.cell(row=row, column=col_idx).value
        if vals.get("lv_nummer") or vals.get("projektname"):
            rows.append(vals)

    wb.close()
    return pd.DataFrame(rows)


class Verifikation:
    """Sammelt Pruefergebnisse und gibt eine Zusammenfassung aus."""

    def __init__(self):
        self.ok = 0
        self.fehler = 0
        self.details = []

    def check(self, bedingung: bool, beschreibung: str):
        if bedingung:
            self.ok += 1
            self.details.append(f"  [OK]     {beschreibung}")
        else:
            self.fehler += 1
            self.details.append(f"  [FEHLER] {beschreibung}")

    def abschnitt(self, titel: str):
        self.details.append(f"\n--- {titel} ---")

    def zusammenfassung(self):
        gesamt = self.ok + self.fehler
        print("\n".join(self.details))
        print(f"\n{'='*60}")
        if self.fehler == 0:
            print(f"  ERGEBNIS: Alle {gesamt} Checks bestanden.")
        else:
            print(f"  ERGEBNIS: {self.ok}/{gesamt} Checks bestanden, {self.fehler} Fehler.")
        print(f"{'='*60}")
        return self.fehler


def pruefe_quelldaten(df: pd.DataFrame, v: Verifikation):
    """Kategorie 1: Vergleicht konsolidierte Werte mit bekannten Sollwerten."""
    v.abschnitt("Quelldaten-Abgleich")

    for lv_nr, erwartet in ERWARTETE_WERTE.items():
        zeile = df[df["lv_nummer"] == lv_nr]

        if zeile.empty:
            v.check(False, f"{lv_nr}: Nicht im konsolidierten Excel gefunden")
            continue

        row = zeile.iloc[0]

        # Ampelstatus
        ist = row.get("ampelstatus", "")
        soll = erwartet["ampelstatus"]
        v.check(ist == soll, f"{lv_nr}: Ampel = {ist} (erwartet: {soll})")

        # PAG
        ist = row.get("pag", "")
        soll = erwartet["pag"]
        v.check(ist == soll, f"{lv_nr}: PAG = {ist} (erwartet: {soll})")

        # Projektphase
        ist = row.get("projektphase", "")
        soll = erwartet["projektphase"]
        v.check(ist == soll, f"{lv_nr}: Phase = {ist} (erwartet: {soll})")

        # Fertigstellungsgrad
        ist = row.get("fertigstellungsgrad", 0)
        soll = erwartet["fertigstellungsgrad"]
        try:
            ist_f = float(ist) if ist is not None else 0.0
        except (ValueError, TypeError):
            ist_f = -1.0
        v.check(abs(ist_f - soll) < 0.01,
                f"{lv_nr}: Fertigstellungsgrad = {ist_f} (erwartet: {soll})")

        # Meilensteine-Anzahl (aus Text: Semikolon-getrennt)
        ms_text = row.get("meilensteine", "") or ""
        if ms_text:
            ms_anzahl = len([m for m in str(ms_text).split(";") if m.strip()])
        else:
            ms_anzahl = 0
        soll = erwartet["meilensteine_anzahl"]
        v.check(ms_anzahl == soll,
                f"{lv_nr}: Meilensteine = {ms_anzahl} (erwartet: {soll})")


def pruefe_normalisierung(df: pd.DataFrame, v: Verifikation):
    """Kategorie 2: Prueft ob alle Werte in den erlaubten Mengen liegen."""
    v.abschnitt("Normalisierung")

    # Ampelwerte
    gueltige_ampel = set(AMPEL_NORM.values())
    ampel_werte = df["ampelstatus"].dropna().unique()
    ungueltig = [w for w in ampel_werte if w not in gueltige_ampel]
    v.check(len(ungueltig) == 0,
            f"Ampelwerte gueltig" + (f" (ungueltig: {ungueltig})" if ungueltig else ""))

    # PAG-Werte
    gueltige_pag = set(PAG_NORM.values())
    pag_werte = df["pag"].dropna().unique()
    ungueltig = [w for w in pag_werte if w not in gueltige_pag]
    v.check(len(ungueltig) == 0,
            f"PAG-Werte gueltig" + (f" (ungueltig: {ungueltig})" if ungueltig else ""))

    # Phasen-Werte
    gueltige_phase = set(PHASE_NORM.values())
    phase_werte = df["projektphase"].dropna().unique()
    ungueltig = [w for w in phase_werte if w not in gueltige_phase]
    v.check(len(ungueltig) == 0,
            f"Phasen-Werte gueltig" + (f" (ungueltig: {ungueltig})" if ungueltig else ""))


def pruefe_datumsfelder(df: pd.DataFrame, v: Verifikation):
    """Kategorie 3: Prueft Datumsfelder auf bekannte Probleme."""
    v.abschnitt("Datumsfelder")

    datum_spalten = ["start", "ende"]
    jahr_fehler = []
    platzhalter_fehler = []

    for spalte in datum_spalten:
        if spalte not in df.columns:
            continue
        for idx, val in df[spalte].items():
            if val is None:
                continue
            lv = df.at[idx, "lv_nummer"]
            # Jahr < 2000 (bekanntes 1905-Problem)
            if isinstance(val, (datetime, date)) and val.year < 2000:
                jahr_fehler.append(f"{lv}/{spalte}: {val}")
            # Platzhalter-Strings
            if isinstance(val, str) and "TT.MM" in val:
                platzhalter_fehler.append(f"{lv}/{spalte}: {val}")

    v.check(len(jahr_fehler) == 0,
            f"Keine Datumswerte mit Jahr < 2000" +
            (f" (gefunden: {jahr_fehler})" if jahr_fehler else ""))

    v.check(len(platzhalter_fehler) == 0,
            f"Keine Platzhalter in Datumsfeldern" +
            (f" (gefunden: {platzhalter_fehler})" if platzhalter_fehler else ""))


def pruefe_budget(df: pd.DataFrame, v: Verifikation):
    """Kategorie 4: Prueft Budgetdaten auf Plausibilitaet."""
    v.abschnitt("Budget/SAP")

    kosten_spalten = [
        "plankosten_2025", "plankosten_2026", "plankosten_2027",
        "istkosten_2025", "istkosten_2026", "istkosten_2027",
        "plankosten_gesamt", "istkosten_gesamt",
    ]

    # Keine negativen Betraege
    negative = []
    for spalte in kosten_spalten:
        if spalte not in df.columns:
            continue
        for idx, val in df[spalte].items():
            if val is None:
                continue
            try:
                if float(val) < 0:
                    lv = df.at[idx, "lv_nummer"]
                    negative.append(f"{lv}/{spalte}: {val}")
            except (ValueError, TypeError):
                pass

    v.check(len(negative) == 0,
            f"Keine negativen Kostenbetraege" +
            (f" (gefunden: {negative})" if negative else ""))

    # Division by Zero bei istkosten_prozent
    if "istkosten_prozent" in df.columns:
        inf_werte = []
        for idx, val in df["istkosten_prozent"].items():
            if val is not None:
                try:
                    f = float(val)
                    if f != f or abs(f) == float("inf"):  # NaN oder Inf
                        lv = df.at[idx, "lv_nummer"]
                        inf_werte.append(f"{lv}: {val}")
                except (ValueError, TypeError):
                    pass
        v.check(len(inf_werte) == 0,
                f"Kein Division-by-Zero bei istkosten_prozent" +
                (f" (gefunden: {inf_werte})" if inf_werte else ""))

    # plankosten_gesamt = Summe Jahreswerte
    summen_fehler = []
    for idx, row in df.iterrows():
        try:
            p25 = float(row.get("plankosten_2025") or 0)
            p26 = float(row.get("plankosten_2026") or 0)
            p27 = float(row.get("plankosten_2027") or 0)
            gesamt = float(row.get("plankosten_gesamt") or 0)
            if abs((p25 + p26 + p27) - gesamt) > 0.01:
                lv = row.get("lv_nummer", "?")
                summen_fehler.append(f"{lv}: {p25}+{p26}+{p27} != {gesamt}")
        except (ValueError, TypeError):
            pass

    v.check(len(summen_fehler) == 0,
            f"plankosten_gesamt = Summe Jahreswerte" +
            (f" (Abweichungen: {summen_fehler})" if summen_fehler else ""))


def pruefe_vollstaendigkeit(df: pd.DataFrame, v: Verifikation):
    """Kategorie 5: Prueft Pflichtfelder und Projektanzahl."""
    v.abschnitt("Vollstaendigkeit")

    # Pflichtfelder
    pflichtfelder = ["lv_nummer", "projektname", "pag", "ampelstatus"]
    leere_felder = []
    for feld in pflichtfelder:
        if feld not in df.columns:
            leere_felder.append(f"Spalte '{feld}' fehlt komplett")
            continue
        for idx, val in df[feld].items():
            if val is None or (isinstance(val, str) and not val.strip()):
                lv = df.at[idx, "lv_nummer"] if feld != "lv_nummer" else f"Zeile {idx+2}"
                leere_felder.append(f"{lv}: {feld} leer")

    v.check(len(leere_felder) == 0,
            f"Pflichtfelder vollstaendig" +
            (f" (leer: {leere_felder})" if leere_felder else ""))

    # Anzahl Projekte = Anzahl PSB-Dateien
    psb_dateien = [f for f in PSB_DIR.glob("PSB_*.xlsx") if not f.name.startswith("~$")]
    n_psb = len(psb_dateien)
    n_konsolidiert = len(df)
    v.check(n_konsolidiert == n_psb,
            f"Projektanzahl: {n_konsolidiert} konsolidiert, {n_psb} PSB-Dateien")


def pruefe_validierung(df: pd.DataFrame, v: Verifikation):
    """Kategorie 6: Prueft ob Validierungsspalte vorhanden ist."""
    v.abschnitt("Validierungsspalte")

    v.check("_validierung" in df.columns,
            f"Spalte '_validierung' vorhanden")


def main():
    konsolidiert_pfad = OUTPUT_REVIEW / "konsolidiert.xlsx"

    print(f"{'='*60}")
    print(f"  Verifikation: {konsolidiert_pfad.name}")
    print(f"{'='*60}")

    if not konsolidiert_pfad.exists():
        print(f"\n  FEHLER: {konsolidiert_pfad} nicht gefunden.")
        print(f"  Bitte zuerst: python 01_konsolidierung.py")
        sys.exit(1)

    df = lade_konsolidiert(konsolidiert_pfad)
    print(f"\n  {len(df)} Projekte geladen.\n")

    v = Verifikation()

    pruefe_quelldaten(df, v)
    pruefe_normalisierung(df, v)
    pruefe_datumsfelder(df, v)
    pruefe_budget(df, v)
    pruefe_vollstaendigkeit(df, v)
    pruefe_validierung(df, v)

    n_fehler = v.zusammenfassung()
    sys.exit(1 if n_fehler > 0 else 0)


if __name__ == "__main__":
    main()
