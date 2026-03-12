"""
Lokaler HTTP-Server fuer das Dashboard.

1. Liest output/review/konsolidiert.xlsx
2. Konvertiert in docs/consolidated.json (Dashboard-Datenformat)
3. Startet HTTP-Server auf Port 8080
4. Oeffnet Browser

Verwendung: python start_dashboard.py
"""

import http.server
import os
import sys
import webbrowser
import json
from pathlib import Path
from datetime import datetime, date

SCRIPT_DIR = Path(__file__).parent
DOCS_DIR = SCRIPT_DIR / "docs"
SOURCE_XLSX = SCRIPT_DIR / "output" / "review" / "konsolidiert.xlsx"
TARGET_JSON = DOCS_DIR / "consolidated.json"
PORT = 8080


def xlsx_to_json():
    """Konvertiert konsolidiert.xlsx in consolidated.json fuer das Dashboard."""
    try:
        import openpyxl
    except ImportError:
        print("FEHLER: openpyxl nicht installiert. Bitte: pip install openpyxl")
        sys.exit(1)

    if not SOURCE_XLSX.exists():
        print(f"FEHLER: {SOURCE_XLSX} nicht gefunden.")
        print("Bitte zuerst: cd prototype && python 01_konsolidierung.py")
        sys.exit(1)

    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb["Konsolidiert"]

    # Header lesen (Zeile 1)
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers.append(str(val).strip())
        else:
            headers.append(f"col_{col}")

    # Daten lesen
    projekte = []
    for row in range(2, ws.max_row + 1):
        vals = {}
        for col_idx, header in enumerate(headers, 1):
            val = ws.cell(row=row, column=col_idx).value
            if isinstance(val, (datetime, date)):
                val = val.isoformat()
            vals[header] = val

        if not vals.get("lv_nummer") and not vals.get("projektname"):
            continue

        projekte.append(vals)

    wb.close()

    output = {
        "meta": {
            "generiert": datetime.now().isoformat(),
            "anzahl_projekte": len(projekte),
            "quelle": SOURCE_XLSX.name,
        },
        "projekte": projekte,
    }

    with open(TARGET_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)

    print(f"  {len(projekte)} Projekte -> {TARGET_JSON.name}")
    return len(projekte)


def main():
    if not DOCS_DIR.exists():
        print(f"FEHLER: {DOCS_DIR} nicht gefunden.")
        sys.exit(1)

    print("Konvertiere konsolidiert.xlsx -> consolidated.json ...")
    n = xlsx_to_json()
    print(f"  {n} Projekte konvertiert.\n")

    os.chdir(DOCS_DIR)
    handler = http.server.SimpleHTTPRequestHandler
    handler.log_message = lambda *args: None

    print(f"Dashboard: http://localhost:{PORT}")
    print("Beenden mit Ctrl+C\n")

    webbrowser.open(f"http://localhost:{PORT}")

    try:
        with http.server.HTTPServer(("", PORT), handler) as httpd:
            httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nServer beendet.")


if __name__ == "__main__":
    main()
