"""
Generiert synthetische Projektstatusberichte (PSB) als Excel-Dateien.

Die Struktur bildet die echten PSB-Vorlagen der VetMedUni Wien ab,
enthaelt aber ausschliesslich fiktive Daten. Dient als Testbasis
fuer das Konsolidierungsscript und den Promptotyping-Workshop.

Bewusst eingebaute Qualitaetsprobleme:
- PSB 3: "Q1" statt standardisiertem Datum
- PSB 4: Fehlender Indikatorwert
- PSB 5: Inkonsistenter Ampelstatus (gruen trotz Verzoegerung)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import random

# Pfade
SCRIPT_DIR = Path(__file__).parent
TEMPLATE_DIR = SCRIPT_DIR.parent / "data" / "templates"
SAMPLE_DIR = SCRIPT_DIR.parent / "data" / "sample"

# Farben fuer Ampelstatus
AMPEL_FILLS = {
    "gruen": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
    "gelb": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
    "rot": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# --- Synthetische Projektdaten ---

PROJEKTE = [
    {
        "id": "LV-2024-001",
        "titel": "Digitalisierung Labormanagement",
        "kurzbeschreibung": "Einfuehrung eines digitalen Laborinformationssystems zur Verwaltung von Proben, Geraeten und Ergebnissen in der Veterinaeranatomie.",
        "kapitel": "Lehre",
        "auftraggeber": "Vizerektorat Lehre",
        "laufzeit_von": "01.01.2024",
        "laufzeit_bis": "31.12.2026",
        "budget_gesamt": 185000,
        "budget_verbraucht": 92000,
        "projektauftrag": "ja",
        "ampel_aktuell": "gruen",
        "indikatoren": [
            {
                "name": "Anteil digitalisierter Laborprotokolle",
                "einheit": "%",
                "zielwert_2024": 30,
                "istwert_2024": 35,
                "zielwert_2025": 60,
                "istwert_2025": None,
                "zielwert_2026": 90,
            },
            {
                "name": "Anzahl geschulter Mitarbeitender",
                "einheit": "Personen",
                "zielwert_2024": 15,
                "istwert_2024": 18,
                "zielwert_2025": 30,
                "istwert_2025": None,
                "zielwert_2026": 45,
            },
        ],
        "massnahmen": [
            {"name": "Pilotphase Anatomie-Labor", "status": "abgeschlossen", "termin": "30.06.2024"},
            {"name": "Rollout weitere Labore", "status": "in Umsetzung", "termin": "31.03.2025"},
            {"name": "Schulungsprogramm Phase 2", "status": "geplant", "termin": "30.09.2025"},
        ],
        "berichtszeitraum": "Q4/2024",
        "kommentar": "Projekt laeuft planmaessig. Pilotphase erfolgreich abgeschlossen.",
    },
    {
        "id": "LV-2024-002",
        "titel": "Curriculumreform Veterinaermedizin",
        "kurzbeschreibung": "Ueberarbeitung des Studienplans mit Fokus auf One-Health-Ansatz und interdisziplinaere Lehre.",
        "kapitel": "Lehre",
        "auftraggeber": "Rektorat",
        "laufzeit_von": "01.03.2024",
        "laufzeit_bis": "28.02.2027",
        "budget_gesamt": 320000,
        "budget_verbraucht": 145000,
        "projektauftrag": "ja",
        "ampel_aktuell": "gelb",
        "indikatoren": [
            {
                "name": "Anteil ueberarbeiteter Module",
                "einheit": "%",
                "zielwert_2024": 20,
                "istwert_2024": 15,
                "zielwert_2025": 50,
                "istwert_2025": None,
                "zielwert_2026": 80,
            },
            {
                "name": "Evaluierungsergebnis Studierende",
                "einheit": "Skala 1-5",
                "zielwert_2024": 3.5,
                "istwert_2024": 3.2,
                "zielwert_2025": 3.8,
                "istwert_2025": None,
                "zielwert_2026": 4.0,
            },
        ],
        "massnahmen": [
            {"name": "Bestandsaufnahme Module", "status": "abgeschlossen", "termin": "30.06.2024"},
            {"name": "Stakeholder-Workshops", "status": "in Umsetzung", "termin": "31.12.2024"},
            {"name": "Pilotierung neue Module", "status": "verzoegert", "termin": "30.06.2025"},
        ],
        "berichtszeitraum": "Q4/2024",
        "kommentar": "Verzoegerung bei Stakeholder-Workshops aufgrund von Terminproblemen. Zielwert 2024 knapp verfehlt.",
    },
    {
        "id": "LV-2023-005",
        "titel": "Forschungsinfrastruktur Biobank",
        "kurzbeschreibung": "Aufbau einer zentralen Biobank fuer veterinaermedizinische Proben mit standardisierten Lagerungs- und Zugriffsprotokollen.",
        "kapitel": "Forschung",
        "auftraggeber": "Vizerektorat Forschung",
        "laufzeit_von": "01.07.2023",
        "laufzeit_bis": "30.06.2026",
        "budget_gesamt": 450000,
        "budget_verbraucht": 310000,
        "projektauftrag": "ja",
        "ampel_aktuell": "gruen",
        "indikatoren": [
            {
                "name": "Anzahl katalogisierter Proben",
                "einheit": "Stueck",
                "zielwert_2024": 5000,
                "istwert_2024": 5200,
                "zielwert_2025": 8000,
                "istwert_2025": None,
                "zielwert_2026": 12000,
            },
        ],
        "massnahmen": [
            {"name": "Raumumbau und Geraeteinstallation", "status": "abgeschlossen", "termin": "31.12.2023"},
            {"name": "Datenbanksystem implementieren", "status": "abgeschlossen", "termin": "30.06.2024"},
            {"name": "Probenuebernahme Altbestand", "status": "in Umsetzung", "termin": "31.03.2025"},
        ],
        # BEWUSSTER QUALITAETSFEHLER: "Q1" statt standardisiertem Datum
        "berichtszeitraum": "Q1",
        "kommentar": "Probenuebernahme laeuft planmaessig.",
    },
    {
        "id": "LV-2024-003",
        "titel": "IT-Sicherheitskonzept",
        "kurzbeschreibung": "Entwicklung und Implementierung eines umfassenden IT-Sicherheitskonzepts gemaess NIS2-Richtlinie.",
        "kapitel": "Infrastruktur",
        "auftraggeber": "Rektorat",
        "laufzeit_von": "01.04.2024",
        "laufzeit_bis": "31.12.2025",
        "budget_gesamt": 280000,
        "budget_verbraucht": 195000,
        "projektauftrag": "nein",
        "ampel_aktuell": "rot",
        "indikatoren": [
            {
                "name": "Anteil umgesetzter Massnahmen",
                "einheit": "%",
                "zielwert_2024": 40,
                "istwert_2024": 25,
                "zielwert_2025": 100,
                "istwert_2025": None,
            },
            {
                "name": "Anzahl Sicherheitsaudits",
                "einheit": "Stueck",
                "zielwert_2024": 2,
                # BEWUSSTER QUALITAETSFEHLER: Fehlender Istwert
                "istwert_2024": None,
                "zielwert_2025": 4,
                "istwert_2025": None,
            },
        ],
        "massnahmen": [
            {"name": "Gap-Analyse", "status": "abgeschlossen", "termin": "30.06.2024"},
            {"name": "Umsetzung technische Massnahmen", "status": "verzoegert", "termin": "31.12.2024"},
            {"name": "Schulungen IT-Personal", "status": "verzoegert", "termin": "31.03.2025"},
        ],
        "berichtszeitraum": "Q4/2024",
        "kommentar": "Erhebliche Verzoegerung wegen Personalengpass in der IT-Abteilung. Externer Dienstleister wird evaluiert. Projektauftrag liegt noch nicht vor.",
    },
    {
        "id": "LV-2024-004",
        "titel": "Nachhaltigkeitsstrategie Campus",
        "kurzbeschreibung": "Erarbeitung und Umsetzung einer Nachhaltigkeitsstrategie fuer den gesamten Campus inkl. Energieeffizienz und Mobilitaetskonzept.",
        "kapitel": "Infrastruktur",
        "auftraggeber": "Vizerektorat Infrastruktur",
        "laufzeit_von": "01.06.2024",
        "laufzeit_bis": "31.12.2026",
        "budget_gesamt": 520000,
        "budget_verbraucht": 78000,
        "projektauftrag": "ja",
        # BEWUSSTER QUALITAETSFEHLER: Ampel gruen trotz Verzoegerung bei Massnahmen
        "ampel_aktuell": "gruen",
        "indikatoren": [
            {
                "name": "CO2-Reduktion gegenueber Basisjahr",
                "einheit": "%",
                "zielwert_2024": 5,
                "istwert_2024": 3,
                "zielwert_2025": 15,
                "istwert_2025": None,
                "zielwert_2026": 25,
            },
            {
                "name": "Anteil erneuerbare Energie",
                "einheit": "%",
                "zielwert_2024": 40,
                "istwert_2024": 38,
                "zielwert_2025": 55,
                "istwert_2025": None,
                "zielwert_2026": 70,
            },
        ],
        "massnahmen": [
            {"name": "Energieaudit durchfuehren", "status": "abgeschlossen", "termin": "30.09.2024"},
            {"name": "Photovoltaik-Ausbau Dachflaechen", "status": "verzoegert", "termin": "30.06.2025"},
            {"name": "Mobilitaetskonzept erarbeiten", "status": "geplant", "termin": "31.12.2025"},
        ],
        "berichtszeitraum": "Q4/2024",
        "kommentar": "Energieaudit abgeschlossen. PV-Ausbau verzoegert wegen Genehmigungsverfahren. Zielwerte knapp verfehlt.",
    },
]


def create_psb_template(wb, ws):
    """Erstellt die leere PSB-Vorlage (Kopfbereich und Struktur)."""

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 40

    # Titel
    ws.merge_cells("A1:H1")
    ws["A1"] = "Projektstatusbericht (PSB)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Metadaten-Block
    meta_fields = [
        ("A3", "Projekt-ID:"),
        ("A4", "Projekttitel:"),
        ("A5", "Kapitel:"),
        ("A6", "Auftraggeber:"),
        ("A7", "Laufzeit von:"),
        ("A8", "Laufzeit bis:"),
        ("A9", "Budget gesamt (EUR):"),
        ("A10", "Budget verbraucht (EUR):"),
        ("A11", "Projektauftrag vorhanden:"),
        ("A12", "Berichtszeitraum:"),
        ("A13", "Ampelstatus:"),
    ]

    for cell_ref, label in meta_fields:
        ws[cell_ref] = label
        ws[cell_ref].font = Font(bold=True)

    return 15  # Startzeile fuer Indikatoren


def fill_metadata(ws, projekt):
    """Fuellt die Metadaten eines Projekts ein."""
    ws["B3"] = projekt["id"]
    ws["B4"] = projekt["titel"]
    ws["B5"] = projekt["kapitel"]
    ws["B6"] = projekt["auftraggeber"]
    ws["B7"] = projekt["laufzeit_von"]
    ws["B8"] = projekt["laufzeit_bis"]
    ws["B9"] = projekt["budget_gesamt"]
    ws["B9"].number_format = '#,##0'
    ws["B10"] = projekt["budget_verbraucht"]
    ws["B10"].number_format = '#,##0'
    ws["B11"] = projekt["projektauftrag"]
    ws["B12"] = projekt["berichtszeitraum"]
    ws["B13"] = projekt["ampel_aktuell"]

    # Ampelfarbe in Zelle
    ampel = projekt["ampel_aktuell"]
    if ampel in AMPEL_FILLS:
        ws["B13"].fill = AMPEL_FILLS[ampel]
        if ampel == "rot":
            ws["B13"].font = Font(color="FFFFFF", bold=True)


def write_indikatoren(ws, projekt, start_row):
    """Schreibt die Indikatorentabelle."""
    row = start_row

    # Sektions-Header
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Indikatoren"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1

    # Tabellen-Header
    headers = ["Indikator", "Einheit", "Zielwert 2024", "Istwert 2024",
               "Zielwert 2025", "Istwert 2025", "Zielwert 2026"]
    cols = ["A", "B", "C", "D", "E", "F", "G"]
    for col, header in zip(cols, headers):
        cell = ws[f"{col}{row}"]
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row += 1

    # Datenzeilen
    for ind in projekt["indikatoren"]:
        ws[f"A{row}"] = ind["name"]
        ws[f"B{row}"] = ind["einheit"]
        ws[f"C{row}"] = ind.get("zielwert_2024")
        ws[f"D{row}"] = ind.get("istwert_2024")
        ws[f"E{row}"] = ind.get("zielwert_2025")
        ws[f"F{row}"] = ind.get("istwert_2025")
        ws[f"G{row}"] = ind.get("zielwert_2026")

        for col in cols:
            ws[f"{col}{row}"].border = THIN_BORDER
            if col in ["C", "D", "E", "F", "G"]:
                ws[f"{col}{row}"].number_format = '#,##0.0'
        row += 1

    return row + 1


def write_massnahmen(ws, projekt, start_row):
    """Schreibt die Massnahmentabelle."""
    row = start_row

    # Sektions-Header
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Massnahmen"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1

    # Tabellen-Header
    headers = ["Massnahme", "Status", "Termin"]
    cols = ["A", "B", "C"]
    for col, header in zip(cols, headers):
        cell = ws[f"{col}{row}"]
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row += 1

    # Datenzeilen
    for m in projekt["massnahmen"]:
        ws[f"A{row}"] = m["name"]
        ws[f"B{row}"] = m["status"]
        ws[f"C{row}"] = m["termin"]
        for col in cols:
            ws[f"{col}{row}"].border = THIN_BORDER
        row += 1

    return row + 1


def write_kommentar(ws, projekt, start_row):
    """Schreibt den Kommentarbereich."""
    row = start_row

    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Kommentar / Erlaeuterung"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1

    ws.merge_cells(f"A{row}:H{row + 2}")
    ws[f"A{row}"] = projekt["kommentar"]
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    return row + 4


def write_kurzbeschreibung(ws, projekt, start_row):
    """Schreibt die Kurzbeschreibung."""
    row = start_row

    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Kurzbeschreibung"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1

    ws.merge_cells(f"A{row}:H{row + 2}")
    ws[f"A{row}"] = projekt["kurzbeschreibung"]
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    return row + 4


def generate_psb(projekt, output_dir):
    """Generiert eine PSB-Excel-Datei fuer ein Projekt."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PSB"

    start_row = create_psb_template(wb, ws)
    fill_metadata(ws, projekt)
    row = write_indikatoren(ws, projekt, start_row)
    row = write_massnahmen(ws, projekt, row)
    row = write_kommentar(ws, projekt, row)
    row = write_kurzbeschreibung(ws, projekt, row)

    filename = f"PSB_{projekt['id']}.xlsx"
    filepath = output_dir / filename
    wb.save(filepath)
    print(f"  Erstellt: {filepath.name}")
    return filepath


def generate_template(template_dir):
    """Generiert eine leere PSB-Vorlage."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PSB"

    start_row = create_psb_template(wb, ws)

    # Leere Indikatoren-Struktur
    row = start_row
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Indikatoren"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1

    headers = ["Indikator", "Einheit", "Zielwert 2024", "Istwert 2024",
               "Zielwert 2025", "Istwert 2025", "Zielwert 2026"]
    cols = ["A", "B", "C", "D", "E", "F", "G"]
    for col, header in zip(cols, headers):
        cell = ws[f"{col}{row}"]
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row += 3  # 2 leere Zeilen fuer Daten

    # Leere Massnahmen-Struktur
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Massnahmen"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT
    row += 1

    headers_m = ["Massnahme", "Status", "Termin"]
    cols_m = ["A", "B", "C"]
    for col, header in zip(cols_m, headers_m):
        cell = ws[f"{col}{row}"]
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row += 3

    # Kommentar
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "Kommentar / Erlaeuterung"
    ws[f"A{row}"].fill = SECTION_FILL
    ws[f"A{row}"].font = SECTION_FONT

    filepath = template_dir / "PSB_Vorlage.xlsx"
    wb.save(filepath)
    print(f"  Vorlage erstellt: {filepath.name}")


def main():
    print("=== Synthetische PSB-Daten generieren ===\n")

    # Verzeichnisse sicherstellen
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    SAMPLE_DIR.mkdir(parents=True, exist_ok=True)

    # Leere Vorlage
    print("Vorlage:")
    generate_template(TEMPLATE_DIR)

    # Synthetische PSBs
    print(f"\n{len(PROJEKTE)} Projektstatusberichte:")
    for projekt in PROJEKTE:
        generate_psb(projekt, SAMPLE_DIR)

    print(f"\nFertig. Dateien in:")
    print(f"  Vorlage:  {TEMPLATE_DIR}")
    print(f"  Testdaten: {SAMPLE_DIR}")

    # Qualitaetsprobleme dokumentieren
    print("\nBewusst eingebaute Qualitaetsprobleme:")
    print("  PSB_LV-2023-005: Berichtszeitraum 'Q1' statt 'Q1/2025'")
    print("  PSB_LV-2024-003: Fehlender Istwert bei 'Anzahl Sicherheitsaudits'")
    print("  PSB_LV-2024-004: Ampel 'gruen' trotz verzoegerter Massnahmen und verfehlter Zielwerte")


if __name__ == "__main__":
    main()
