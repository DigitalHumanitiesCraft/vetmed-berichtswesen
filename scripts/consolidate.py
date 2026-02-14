"""
Konsolidierungsscript fuer Projektstatusberichte (PSB).

Liest alle PSB-Excel-Dateien aus einem Verzeichnis, extrahiert die
strukturierten Daten und fuehrt sie in eine konsolidierte Uebersicht
zusammen. Entspricht Stufe 2 des 4-Stufen-Workflows:
Erfassung -> **Konsolidierung** -> Dashboard -> Export.

Das Script enthaelt bewusst eine manuelle Ueberpruefungsphase:
Auffaelligkeiten und Qualitaetsprobleme werden in einem separaten
Report dokumentiert, damit der Fachverantwortliche sie pruefen kann.
"""

import openpyxl
import json
import csv
import sys
from pathlib import Path
from datetime import datetime

SCRIPT_DIR = Path(__file__).parent
DEFAULT_INPUT = SCRIPT_DIR.parent / "data" / "sample"
OUTPUT_DIR = SCRIPT_DIR.parent / "data" / "consolidated"

# Erwartetes Format fuer Berichtszeitraum: "Q1/2024", "Q2/2025" etc.
VALID_QUARTER_PATTERN = r"^Q[1-4]/\d{4}$"


def parse_psb(filepath):
    """
    Extrahiert strukturierte Daten aus einer PSB-Excel-Datei.

    Returns:
        dict mit Metadaten, Indikatoren, Massnahmen, Kommentar
        oder None bei Fehler
    """
    import re

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["PSB"]

    projekt = {
        "quelldatei": filepath.name,
        "id": ws["B3"].value,
        "titel": ws["B4"].value,
        "kapitel": ws["B5"].value,
        "auftraggeber": ws["B6"].value,
        "laufzeit_von": ws["B7"].value,
        "laufzeit_bis": ws["B8"].value,
        "budget_gesamt": ws["B9"].value,
        "budget_verbraucht": ws["B10"].value,
        "projektauftrag": ws["B11"].value,
        "berichtszeitraum": ws["B12"].value,
        "ampel": ws["B13"].value,
        "indikatoren": [],
        "massnahmen": [],
        "kommentar": None,
        "kurzbeschreibung": None,
        "warnungen": [],
    }

    # --- Qualitaetspruefungen ---

    # Berichtszeitraum validieren
    bz = str(projekt["berichtszeitraum"] or "")
    if not re.match(VALID_QUARTER_PATTERN, bz):
        projekt["warnungen"].append(
            f"Berichtszeitraum '{bz}' entspricht nicht dem Format 'Qn/YYYY'"
        )

    # Projektauftrag pruefen
    if projekt["projektauftrag"] and projekt["projektauftrag"].lower() == "nein":
        projekt["warnungen"].append("Kein Projektauftrag vorhanden")

    # Budget-Konsistenz
    if projekt["budget_gesamt"] and projekt["budget_verbraucht"]:
        if projekt["budget_verbraucht"] > projekt["budget_gesamt"]:
            projekt["warnungen"].append(
                f"Budget ueberschritten: {projekt['budget_verbraucht']} > {projekt['budget_gesamt']}"
            )
        verbrauch_pct = projekt["budget_verbraucht"] / projekt["budget_gesamt"] * 100
        projekt["budget_verbrauch_pct"] = round(verbrauch_pct, 1)

    # --- Indikatoren extrahieren ---
    # Suche nach dem Indikator-Header
    ind_start = None
    for row in range(15, 50):
        val = ws[f"A{row}"].value
        if val and "Indikator" in str(val) and ws[f"B{row}"].value == "Einheit":
            ind_start = row + 1
            break

    if ind_start:
        row = ind_start
        while row < ind_start + 10:
            name = ws[f"A{row}"].value
            if not name or ("Massnahme" in str(name) and ws[f"B{row}"].value == "Status"):
                break
            indikator = {
                "name": name,
                "einheit": ws[f"B{row}"].value,
                "zielwert_2024": ws[f"C{row}"].value,
                "istwert_2024": ws[f"D{row}"].value,
                "zielwert_2025": ws[f"E{row}"].value,
                "istwert_2025": ws[f"F{row}"].value,
                "zielwert_2026": ws[f"G{row}"].value,
            }

            # Fehlende Istwerte pruefen (fuer abgelaufene Zeitraeume)
            if indikator["zielwert_2024"] is not None and indikator["istwert_2024"] is None:
                projekt["warnungen"].append(
                    f"Indikator '{name}': Istwert 2024 fehlt"
                )

            # Zielverfehlung pruefen
            if (indikator["zielwert_2024"] is not None
                    and indikator["istwert_2024"] is not None):
                try:
                    ziel = float(indikator["zielwert_2024"])
                    ist = float(indikator["istwert_2024"])
                    if ist < ziel:
                        abweichung = round((ist - ziel) / ziel * 100, 1)
                        projekt["warnungen"].append(
                            f"Indikator '{name}': Zielwert 2024 verfehlt ({abweichung}%)"
                        )
                except (ValueError, TypeError):
                    pass

            projekt["indikatoren"].append(indikator)
            row += 1

    # --- Massnahmen extrahieren ---
    mass_start = None
    for row in range(ind_start or 15, 60):
        val = ws[f"A{row}"].value
        if val and "Massnahme" in str(val) and ws[f"B{row}"].value == "Status":
            mass_start = row + 1
            break

    if mass_start:
        row = mass_start
        while row < mass_start + 10:
            name = ws[f"A{row}"].value
            if not name or "Kommentar" in str(name):
                break
            massnahme = {
                "name": name,
                "status": ws[f"B{row}"].value,
                "termin": ws[f"C{row}"].value,
            }

            # Verzoegerte Massnahmen bei gruener Ampel
            if (massnahme["status"] and "verzoegert" in str(massnahme["status"]).lower()
                    and projekt["ampel"] and projekt["ampel"].lower() == "gruen"):
                projekt["warnungen"].append(
                    f"Massnahme '{name}' verzoegert, aber Ampel ist gruen"
                )

            projekt["massnahmen"].append(massnahme)
            row += 1

    # --- Kommentar extrahieren ---
    for row in range(mass_start or 20, 70):
        val = ws[f"A{row}"].value
        if val and "Kommentar" in str(val):
            kommentar_row = row + 1
            projekt["kommentar"] = ws[f"A{kommentar_row}"].value
            break

    # --- Kurzbeschreibung extrahieren ---
    for row in range(mass_start or 20, 70):
        val = ws[f"A{row}"].value
        if val and "Kurzbeschreibung" in str(val):
            kb_row = row + 1
            projekt["kurzbeschreibung"] = ws[f"A{kb_row}"].value
            break

    wb.close()
    return projekt


def consolidate(input_dir, output_dir):
    """
    Konsolidiert alle PSB-Dateien aus input_dir.

    Erzeugt:
    - consolidated.json: Alle Daten als strukturiertes JSON
    - consolidated.csv: Uebersichtstabelle (flach)
    - quality_report.md: Bericht ueber Auffaelligkeiten
    """
    input_path = Path(input_dir)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    psb_files = sorted(input_path.glob("PSB_*.xlsx"))
    if not psb_files:
        print(f"Keine PSB-Dateien in {input_path} gefunden.")
        return

    print(f"Konsolidiere {len(psb_files)} PSB-Dateien...\n")

    projekte = []
    alle_warnungen = []

    for f in psb_files:
        print(f"  Verarbeite: {f.name}")
        projekt = parse_psb(f)
        if projekt:
            projekte.append(projekt)
            if projekt["warnungen"]:
                alle_warnungen.append((projekt["id"], projekt["titel"], projekt["warnungen"]))

    # --- JSON-Export ---
    json_path = output_path / "consolidated.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "meta": {
                    "generiert": datetime.now().isoformat(),
                    "anzahl_projekte": len(projekte),
                    "quelldateien": [p["quelldatei"] for p in projekte],
                },
                "projekte": projekte,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )
    print(f"\n  JSON: {json_path.name}")

    # --- CSV-Export (flache Uebersicht) ---
    csv_path = output_path / "consolidated.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow([
            "ID", "Titel", "Kapitel", "Auftraggeber", "Ampel",
            "Budget gesamt", "Budget verbraucht", "Verbrauch %",
            "Laufzeit von", "Laufzeit bis", "Berichtszeitraum",
            "Projektauftrag", "Anzahl Warnungen"
        ])
        for p in projekte:
            writer.writerow([
                p["id"], p["titel"], p["kapitel"], p["auftraggeber"],
                p["ampel"], p["budget_gesamt"], p["budget_verbraucht"],
                p.get("budget_verbrauch_pct", ""),
                p["laufzeit_von"], p["laufzeit_bis"], p["berichtszeitraum"],
                p["projektauftrag"], len(p["warnungen"])
            ])
    print(f"  CSV:  {csv_path.name}")

    # --- Quality Report ---
    report_path = output_path / "quality_report.md"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("# Qualitaetsbericht Konsolidierung\n\n")
        f.write(f"Generiert: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n")
        f.write(f"Verarbeitete PSBs: {len(projekte)}\n\n")

        # Ampel-Uebersicht
        ampel_counts = {}
        for p in projekte:
            a = p["ampel"] or "unbekannt"
            ampel_counts[a] = ampel_counts.get(a, 0) + 1
        f.write("## Ampelstatus-Uebersicht\n\n")
        for status, count in sorted(ampel_counts.items()):
            f.write(f"- **{status}**: {count} Projekt(e)\n")
        f.write("\n")

        # Budget-Uebersicht
        total_budget = sum(p["budget_gesamt"] or 0 for p in projekte)
        total_verbraucht = sum(p["budget_verbraucht"] or 0 for p in projekte)
        f.write("## Budget-Uebersicht\n\n")
        f.write(f"- Gesamtbudget: {total_budget:,.0f} EUR\n")
        f.write(f"- Verbraucht: {total_verbraucht:,.0f} EUR ({total_verbraucht/total_budget*100:.1f}%)\n\n")

        # Warnungen
        if alle_warnungen:
            f.write("## Auffaelligkeiten zur manuellen Pruefung\n\n")
            for proj_id, titel, warnungen in alle_warnungen:
                f.write(f"### {proj_id}: {titel}\n\n")
                for w in warnungen:
                    f.write(f"- {w}\n")
                f.write("\n")
        else:
            f.write("## Auffaelligkeiten\n\nKeine Auffaelligkeiten gefunden.\n")

        # Rote Ampel separat
        rote = [p for p in projekte if p["ampel"] and p["ampel"].lower() == "rot"]
        if rote:
            f.write("## Projekte mit roter Ampel (gesonderte Vorlage)\n\n")
            for p in rote:
                f.write(f"### {p['id']}: {p['titel']}\n\n")
                f.write(f"- Auftraggeber: {p['auftraggeber']}\n")
                f.write(f"- Budget: {p['budget_verbraucht']:,.0f} / {p['budget_gesamt']:,.0f} EUR\n")
                f.write(f"- Kommentar: {p['kommentar']}\n\n")

    print(f"  Report: {report_path.name}")

    # --- Zusammenfassung ---
    print(f"\n=== Zusammenfassung ===")
    print(f"  Projekte: {len(projekte)}")
    print(f"  Ampelstatus: {ampel_counts}")
    print(f"  Warnungen: {sum(len(w) for _, _, w in alle_warnungen)}")
    if alle_warnungen:
        print(f"\n  ACHTUNG: {len(alle_warnungen)} Projekt(e) mit Auffaelligkeiten.")
        print(f"  Bitte {report_path.name} pruefen und manuell freigeben.")


def main():
    input_dir = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT
    output_dir = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_DIR
    consolidate(input_dir, output_dir)


if __name__ == "__main__":
    main()
